"""
Export Handler Module
Export test cases to JSON, Excel, or other formats
"""

import json
from typing import List, Dict, Any
from io import BytesIO
from pathlib import Path


class ExportHandler:
    """Export test cases in various formats"""
    
    @staticmethod
    def to_json(test_cases: List['TestCase'], pretty: bool = True) -> str:
        """
        Export test cases to JSON
        
        Args:
            test_cases: List of test cases
            pretty: Pretty print JSON
            
        Returns:
            JSON string
        """
        data = [tc.to_dict() for tc in test_cases]
        
        if pretty:
            return json.dumps(data, indent=2, ensure_ascii=False)
        else:
            return json.dumps(data, ensure_ascii=False)
    
    @staticmethod
    def to_json_file(test_cases: List['TestCase'], filepath: str) -> None:
        """
        Export test cases to JSON file
        
        Args:
            test_cases: List of test cases
            filepath: Output file path
        """
        json_str = ExportHandler.to_json(test_cases, pretty=True)
        
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(json_str)
    
    @staticmethod
    def to_excel(test_cases: List['TestCase'], filepath: str) -> None:
        """
        Export test cases to Excel file
        
        Args:
            test_cases: List of test cases
            filepath: Output file path
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Header
        headers = [
            "Test ID", "Requirement ID", "Title", "Description",
            "Preconditions", "Steps", "Expected Result",
            "Type", "Priority", "Domain"
        ]
        
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True)
        
        # Add data rows
        for tc in test_cases:
            row = [
                tc.test_id,
                tc.requirement_id,
                tc.title,
                tc.description,
                "; ".join(tc.preconditions),
                "; ".join(tc.steps),
                tc.expected_result,
                tc.test_type,
                tc.priority,
                tc.domain,
            ]
            ws.append(row)
            
            # Wrap text for all cells
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 30
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 12
        ws.column_dimensions["J"].width = 15
        
        wb.save(filepath)
    
    @staticmethod
    def to_excel_bytes(test_cases: List['TestCase']) -> bytes:
        """
        Export test cases to Excel and return as bytes
        (for FastAPI file response)
        
        Args:
            test_cases: List of test cases
            
        Returns:
            Excel file as bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Header
        headers = [
            "Test ID", "Requirement ID", "Title", "Description",
            "Preconditions", "Steps", "Expected Result",
            "Type", "Priority", "Domain"
        ]
        
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data rows
        for tc in test_cases:
            row = [
                tc.test_id,
                tc.requirement_id,
                tc.title,
                tc.description,
                "; ".join(tc.preconditions),
                "; ".join(tc.steps),
                tc.expected_result,
                tc.test_type,
                tc.priority,
                tc.domain,
            ]
            ws.append(row)
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def to_csv(test_cases: List['TestCase'], filepath: str) -> None:
        """
        Export test cases to CSV
        
        Args:
            test_cases: List of test cases
            filepath: Output file path
        """
        import csv
        
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([
                "Test ID", "Requirement ID", "Title", "Description",
                "Preconditions", "Steps", "Expected Result",
                "Type", "Priority", "Domain"
            ])
            
            # Data
            for tc in test_cases:
                writer.writerow([
                    tc.test_id,
                    tc.requirement_id,
                    tc.title,
                    tc.description,
                    "; ".join(tc.preconditions),
                    "; ".join(tc.steps),
                    tc.expected_result,
                    tc.test_type,
                    tc.priority,
                    tc.domain,
                ])
    
    @staticmethod
    def to_markdown(test_cases: List['TestCase']) -> str:
        """
        Export test cases to Markdown format
        
        Args:
            test_cases: List of test cases
            
        Returns:
            Markdown string
        """
        md = "# Test Cases\n\n"
        md += f"**Total: {len(test_cases)} test cases**\n\n"
        
        # Group by requirement
        by_req = {}
        for tc in test_cases:
            if tc.requirement_id not in by_req:
                by_req[tc.requirement_id] = []
            by_req[tc.requirement_id].append(tc)
        
        for req_id, tests in sorted(by_req.items()):
            md += f"## {req_id}\n\n"
            
            for tc in tests:
                md += f"### {tc.test_id}: {tc.title}\n\n"
                md += f"- **Type:** {tc.test_type}\n"
                md += f"- **Priority:** {tc.priority}\n"
                md += f"- **Domain:** {tc.domain}\n\n"
                
                md += "**Description:** " + tc.description + "\n\n"
                
                md += "**Preconditions:**\n"
                for pre in tc.preconditions:
                    md += f"- {pre}\n"
                md += "\n"
                
                md += "**Steps:**\n"
                for i, step in enumerate(tc.steps, 1):
                    md += f"{i}. {step}\n"
                md += "\n"
                
                md += f"**Expected Result:** {tc.expected_result}\n\n"
                md += "---\n\n"
        
        return md
    
    @staticmethod
    def to_markdown_file(test_cases: List['TestCase'], filepath: str) -> None:
        """
        Export test cases to Markdown file
        
        Args:
            test_cases: List of test cases
            filepath: Output file path
        """
        md = ExportHandler.to_markdown(test_cases)
        
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(md)
