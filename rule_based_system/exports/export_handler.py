"""
Export handler module.
Multi-format export: JSON, Excel, CSV, Markdown.
"""

import json
import os
from typing import List
from ..models.canonical import TestCase


def export_json(test_cases: List[TestCase], output_path: str = None) -> str:
    """
    Export test cases to JSON format.
    
    Args:
        test_cases: List of TestCase objects
        output_path: Path to save JSON file. If None, return JSON string
        
    Returns:
        JSON string or filepath if output_path provided
    """
    data = [tc.to_dict() for tc in test_cases]
    json_str = json.dumps(data, indent=2, ensure_ascii=False)
    
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(json_str)
        return output_path
    
    return json_str


def export_csv(test_cases: List[TestCase], output_path: str) -> str:
    """
    Export test cases to CSV format.
    
    Args:
        test_cases: List of TestCase objects
        output_path: Path to save CSV file
        
    Returns:
        Filepath
    """
    try:
        import csv
    except ImportError:
        raise ImportError("CSV module is built-in, but something went wrong")
    
    headers = ["id", "req_id", "title", "precondition", "steps", 
               "expected_result", "test_type", "priority"]
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        
        for tc in test_cases:
            writer.writerow(tc.to_dict())
    
    return output_path


def export_excel(test_cases: List[TestCase], output_path: str) -> str:
    """
    Export test cases to Excel format.
    
    Args:
        test_cases: List of TestCase objects
        output_path: Path to save XLSX file
        
    Returns:
        Filepath
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        raise ImportError("Cần cài: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    headers = ["ID", "Req ID", "Title", "Precondition", "Steps", 
               "Expected Result", "Type", "Priority"]
    
    # Add header row with styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for row_idx, tc in enumerate(test_cases, 2):
        data = tc.to_dict()
        ws.cell(row=row_idx, column=1).value = data["id"]
        ws.cell(row=row_idx, column=2).value = data["req_id"]
        ws.cell(row=row_idx, column=3).value = data["title"]
        ws.cell(row=row_idx, column=4).value = data["precondition"]
        ws.cell(row=row_idx, column=5).value = data["steps"]
        ws.cell(row=row_idx, column=6).value = data["expected_result"]
        ws.cell(row=row_idx, column=7).value = data["test_type"]
        ws.cell(row=row_idx, column=8).value = data["priority"]
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    wb.save(output_path)
    return output_path


def export_markdown(test_cases: List[TestCase], output_path: str = None) -> str:
    """
    Export test cases to Markdown format.
    
    Args:
        test_cases: List of TestCase objects
        output_path: Path to save Markdown file. If None, return markdown string
        
    Returns:
        Markdown string or filepath if output_path provided
    """
    md_lines = [
        "# Test Cases Report\n",
        f"**Total Test Cases:** {len(test_cases)}\n",
        "\n---\n",
    ]
    
    # Group by requirement ID
    grouped = {}
    for tc in test_cases:
        if tc.req_id not in grouped:
            grouped[tc.req_id] = []
        grouped[tc.req_id].append(tc)
    
    # Generate markdown for each requirement group
    for req_id, req_tests in grouped.items():
        md_lines.append(f"## Requirement: {req_id}\n\n")
        
        for tc in req_tests:
            md_lines.append(f"### {tc.title}\n")
            md_lines.append(f"- **Test ID:** {tc.id}\n")
            md_lines.append(f"- **Type:** {tc.test_type}\n")
            md_lines.append(f"- **Priority:** {tc.priority}\n")
            md_lines.append(f"- **Precondition:** {tc.precondition}\n")
            md_lines.append(f"- **Steps:**\n")
            
            for step in tc.steps:
                md_lines.append(f"  1. {step}\n")
            
            md_lines.append(f"- **Expected Result:** {tc.expected_result}\n\n")
        
        md_lines.append("\n---\n\n")
    
    md_str = "".join(md_lines)
    
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(md_str)
        return output_path
    
    return md_str


def export_all_formats(test_cases: List[TestCase], base_path: str) -> dict:
    """
    Export test cases to all supported formats.
    
    Args:
        test_cases: List of TestCase objects
        base_path: Base directory/filename (without extension)
        
    Returns:
        Dict with format names and their output paths
    """
    base_dir = os.path.dirname(base_path)
    base_name = os.path.basename(base_path)
    
    # Create directory if needed
    if base_dir and not os.path.exists(base_dir):
        os.makedirs(base_dir)
    
    results = {
        "json": export_json(test_cases, f"{base_path}.json"),
        "csv": export_csv(test_cases, f"{base_path}.csv"),
        "excel": export_excel(test_cases, f"{base_path}.xlsx"),
        "markdown": export_markdown(test_cases, f"{base_path}.md"),
    }
    
    return results
